from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
sheet = wb.active
sheet.title = "Objects"

# Headers with formatting
headers = ["ID", "Name", "Side", "Type", "X_Coord", "Y_Coord", "Speed", "Direction", 
           "HP", "Max_HP", "Range", "Attack_Power", "Accuracy", "Armor", "Personnel_Count"]

for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Column widths
widths = [8, 25, 10, 15, 12, 12, 10, 12, 10, 12, 10, 15, 12, 10, 18]
for col, width in enumerate(widths, start=1):
    sheet.column_dimensions[chr(64+col)].width = width

# Side A (Blue) - starting at 30.0, 50.0
data_a = [
    # Tanks
    [1, "T-72B3 Tank #1", "A", "tank", 30.05, 50.05, 0.002, 90, 100, 100, 2.5, 80, 0.75, 80, 3],
    [2, "T-72B3 Tank #2", "A", "tank", 30.10, 50.10, 0.002, 90, 100, 100, 2.5, 80, 0.75, 80, 3],
    [3, "T-80BVM Tank #1", "A", "tank", 30.15, 50.15, 0.0025, 90, 110, 110, 2.8, 85, 0.78, 85, 3],
    
    # BMP (Infantry Fighting Vehicles)
    [4, "BMP-2 #1", "A", "bmp", 30.08, 50.20, 0.003, 90, 60, 60, 1.5, 40, 0.70, 30, 7],
    [5, "BMP-2 #2", "A", "bmp", 30.13, 50.25, 0.003, 90, 60, 60, 1.5, 40, 0.70, 30, 7],
    [6, "BTR-80 #1", "A", "bmp", 30.18, 50.30, 0.0035, 90, 50, 50, 1.2, 25, 0.65, 20, 8],
    
    # Infantry
    [7, "Infantry Squad #1", "A", "infantry", 30.06, 50.08, 0.0015, 90, 100, 100, 0.4, 15, 0.60, 5, 10],
    [8, "Infantry Squad #2", "A", "infantry", 30.11, 50.13, 0.0015, 90, 100, 100, 0.4, 15, 0.60, 5, 10],
    [9, "Infantry Squad #3", "A", "infantry", 30.16, 50.18, 0.0015, 90, 100, 100, 0.4, 15, 0.60, 5, 10],
    
    # Mortars
    [10, "2B11 Mortar #1", "A", "mortar", 30.02, 50.02, 0.0005, 90, 40, 40, 5.0, 60, 0.50, 5, 4],
    [11, "2B11 Mortar #2", "A", "mortar", 30.04, 50.04, 0.0005, 90, 40, 40, 5.0, 60, 0.50, 5, 4],
    
    # Artillery
    [12, "2S19 Msta-S #1", "A", "artillery", 30.00, 50.00, 0.0008, 90, 80, 80, 15.0, 120, 0.65, 40, 5],
    [13, "2S19 Msta-S #2", "A", "artillery", 30.01, 50.01, 0.0008, 90, 80, 80, 15.0, 120, 0.65, 40, 5],
    
    # UAVs
    [14, "Orlan-10 UAV #1", "A", "uav", 30.25, 50.25, 0.008, 90, 20, 20, 10.0, 5, 0.90, 1, 0],
    [15, "Orlan-10 UAV #2", "A", "uav", 30.30, 50.30, 0.008, 90, 20, 20, 10.0, 5, 0.90, 1, 0],
]

# Side B (Red) - starting at 30.5, 50.5
data_b = [
    # Tanks
    [16, "M1A2 Abrams #1", "B", "tank", 30.55, 50.55, 0.0022, 270, 120, 120, 2.8, 90, 0.80, 90, 3],
    [17, "M1A2 Abrams #2", "B", "tank", 30.50, 50.50, 0.0022, 270, 120, 120, 2.8, 90, 0.80, 90, 3],
    [18, "M1A2 Abrams #3", "B", "tank", 30.45, 50.45, 0.0022, 270, 120, 120, 2.8, 90, 0.80, 90, 3],
    
    # BMP
    [19, "M2 Bradley #1", "B", "bmp", 30.52, 50.40, 0.0032, 270, 65, 65, 1.8, 45, 0.73, 35, 6],
    [20, "M2 Bradley #2", "B", "bmp", 30.47, 50.35, 0.0032, 270, 65, 65, 1.8, 45, 0.73, 35, 6],
    [21, "Stryker #1", "B", "bmp", 30.42, 50.30, 0.0038, 270, 55, 55, 1.5, 30, 0.68, 25, 7],
    
    # Infantry
    [22, "Infantry Squad #1", "B", "infantry", 30.54, 50.52, 0.0015, 270, 100, 100, 0.4, 15, 0.60, 5, 10],
    [23, "Infantry Squad #2", "B", "infantry", 30.49, 50.47, 0.0015, 270, 100, 100, 0.4, 15, 0.60, 5, 10],
    [24, "Infantry Squad #3", "B", "infantry", 30.44, 50.42, 0.0015, 270, 100, 100, 0.4, 15, 0.60, 5, 10],
    
    # Mortars
    [25, "M252 Mortar #1", "B", "mortar", 30.58, 50.58, 0.0005, 270, 40, 40, 5.5, 65, 0.55, 5, 4],
    [26, "M252 Mortar #2", "B", "mortar", 30.56, 50.56, 0.0005, 270, 40, 40, 5.5, 65, 0.55, 5, 4],
    
    # Artillery
    [27, "M109A7 Paladin #1", "B", "artillery", 30.60, 50.60, 0.0008, 270, 85, 85, 18.0, 130, 0.68, 45, 5],
    [28, "M109A7 Paladin #2", "B", "artillery", 30.62, 50.62, 0.0008, 270, 85, 85, 18.0, 130, 0.68, 45, 5],
    
    # UAVs
    [29, "RQ-11 Raven #1", "B", "uav", 30.35, 50.35, 0.009, 270, 15, 15, 8.0, 0, 0.95, 1, 0],
    [30, "MQ-9 Reaper #1", "B", "uav", 30.40, 50.40, 0.010, 270, 30, 30, 12.0, 80, 0.85, 5, 0],
]

# Add all data
for row_data in data_a + data_b:
    sheet.append(row_data)

# Format coordinates as numbers with 5 decimals
for row in range(2, sheet.max_row + 1):
    for col in [5, 6]:  # X_Coord, Y_Coord
        sheet.cell(row=row, column=col).number_format = '0.00000'
    
    # Format other numbers
    sheet.cell(row=row, column=7).number_format = '0.0000'  # Speed
    sheet.cell(row=row, column=8).number_format = '0'       # Direction
    sheet.cell(row=row, column=11).number_format = '0.0'    # Range
    sheet.cell(row=row, column=13).number_format = '0.00'   # Accuracy

# Add notes sheet with descriptions
notes_sheet = wb.create_sheet("Notes")
notes_sheet['A1'] = "Field Descriptions"
notes_sheet['A1'].font = Font(bold=True, size=14)

descriptions = [
    ["", ""],
    ["Field", "Description"],
    ["ID", "Unique identifier for each unit"],
    ["Name", "Display name of the unit"],
    ["Side", "A or B - opposing sides in conflict"],
    ["Type", "Unit type: tank, bmp, infantry, mortar, artillery, uav"],
    ["X_Coord", "Longitude coordinate (decimal degrees)"],
    ["Y_Coord", "Latitude coordinate (decimal degrees)"],
    ["Speed", "Movement speed (degrees per step)"],
    ["Direction", "Initial direction in degrees (0-360)"],
    ["HP", "Current health points"],
    ["Max_HP", "Maximum health points"],
    ["Range", "Attack range in km"],
    ["Attack_Power", "Base damage output"],
    ["Accuracy", "Hit probability (0.0-1.0)"],
    ["Armor", "Damage reduction"],
    ["Personnel_Count", "Number of personnel (0 for unmanned)"],
]

for row_idx, row_data in enumerate(descriptions, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = notes_sheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx == 2:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

notes_sheet.column_dimensions['A'].width = 18
notes_sheet.column_dimensions['B'].width = 55

wb.save('/home/claude/combat_simulation/data/objects.xlsx')
print("objects.xlsx created successfully")
