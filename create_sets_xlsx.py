from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "Engagement_Rules"

# Headers
headers = ["Attacker_Type", "Target_Type", "Base_Hit_Probability", "Damage_Multiplier", 
           "Min_Range", "Max_Range", "Engagement_Priority", "Notes"]

for col, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Column widths
widths = [18, 15, 22, 20, 12, 12, 20, 40]
for col, width in enumerate(widths, start=1):
    col_letter = chr(64+col) if col <= 26 else f"A{chr(64+col-26)}"
    sheet.column_dimensions[col_letter].width = width

# Engagement rules data
rules = [
    # Tank engagements
    ["tank", "tank", 0.75, 1.2, 0.5, 2.5, 1, "Direct fire, armor-piercing rounds"],
    ["tank", "bmp", 0.85, 1.5, 0.5, 2.5, 2, "High effectiveness against light armor"],
    ["tank", "infantry", 0.60, 0.8, 0.2, 1.5, 4, "Anti-personnel rounds less effective"],
    ["tank", "mortar", 0.70, 1.3, 0.5, 2.5, 3, "Direct fire on artillery positions"],
    ["tank", "artillery", 0.70, 1.3, 0.5, 2.5, 3, "Direct fire on artillery positions"],
    ["tank", "uav", 0.20, 0.5, 0.1, 1.0, 5, "Difficult to engage aerial targets"],
    
    # BMP engagements
    ["bmp", "tank", 0.35, 0.6, 0.3, 1.5, 5, "Limited effectiveness vs heavy armor"],
    ["bmp", "bmp", 0.70, 1.0, 0.3, 1.8, 2, "Equal combat capability"],
    ["bmp", "infantry", 0.80, 1.2, 0.2, 1.5, 1, "Autocannon effective vs infantry"],
    ["bmp", "mortar", 0.65, 1.1, 0.3, 1.8, 3, "Good fire support suppression"],
    ["bmp", "artillery", 0.60, 1.0, 0.3, 1.8, 3, "Mobile counter-battery"],
    ["bmp", "uav", 0.25, 0.6, 0.1, 1.2, 4, "Some AA capability"],
    
    # Infantry engagements
    ["infantry", "tank", 0.25, 0.4, 0.1, 0.4, 4, "RPG-7 close range only"],
    ["infantry", "bmp", 0.40, 0.7, 0.1, 0.4, 3, "RPG effective at close range"],
    ["infantry", "infantry", 0.65, 1.0, 0.05, 0.4, 1, "Small arms, close combat"],
    ["infantry", "mortar", 0.50, 0.8, 0.05, 0.3, 2, "Overrun positions"],
    ["infantry", "artillery", 0.45, 0.7, 0.05, 0.3, 2, "Vulnerable at close range"],
    ["infantry", "uav", 0.15, 0.3, 0.05, 0.2, 5, "Small arms fire, low probability"],
    
    # Mortar engagements
    ["mortar", "tank", 0.30, 0.5, 1.0, 5.0, 4, "Indirect fire, top attack"],
    ["mortar", "bmp", 0.40, 0.7, 1.0, 5.0, 3, "Area suppression"],
    ["mortar", "infantry", 0.70, 1.5, 1.0, 5.0, 1, "High effectiveness vs soft targets"],
    ["mortar", "mortar", 0.50, 1.0, 1.0, 5.0, 2, "Counter-battery fire"],
    ["mortar", "artillery", 0.45, 0.9, 1.0, 5.0, 2, "Counter-battery fire"],
    ["mortar", "uav", 0.05, 0.2, 0.0, 0.0, 5, "Cannot engage aerial targets"],
    
    # Artillery engagements
    ["artillery", "tank", 0.40, 0.8, 3.0, 18.0, 3, "Precision strike capability"],
    ["artillery", "bmp", 0.50, 1.0, 3.0, 18.0, 3, "Area denial"],
    ["artillery", "infantry", 0.75, 2.0, 3.0, 18.0, 2, "Devastating vs infantry"],
    ["artillery", "mortar", 0.60, 1.3, 3.0, 18.0, 1, "Counter-battery primary mission"],
    ["artillery", "artillery", 0.65, 1.2, 3.0, 18.0, 1, "Counter-battery primary mission"],
    ["artillery", "uav", 0.05, 0.2, 0.0, 0.0, 5, "Cannot engage aerial targets"],
    
    # UAV engagements (reconnaissance and strike)
    ["uav", "tank", 0.60, 0.9, 0.5, 12.0, 2, "Armed UAV precision strike"],
    ["uav", "bmp", 0.65, 1.0, 0.5, 12.0, 2, "Effective vs light armor"],
    ["uav", "infantry", 0.55, 0.8, 0.5, 12.0, 3, "Small target, harder to hit"],
    ["uav", "mortar", 0.70, 1.2, 0.5, 12.0, 1, "Priority target for armed UAV"],
    ["uav", "artillery", 0.70, 1.2, 0.5, 12.0, 1, "Priority target for armed UAV"],
    ["uav", "uav", 0.20, 0.5, 0.5, 8.0, 4, "Air-to-air very limited"],
]

# Add data
for row_data in rules:
    sheet.append(row_data)

# Format probability and multiplier columns
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=3).number_format = '0.00'  # Base_Hit_Probability
    sheet.cell(row=row, column=4).number_format = '0.0'   # Damage_Multiplier
    sheet.cell(row=row, column=5).number_format = '0.0'   # Min_Range
    sheet.cell(row=row, column=6).number_format = '0.0'   # Max_Range

# Add modifiers sheet
mod_sheet = wb.create_sheet("Combat_Modifiers")
mod_sheet['A1'] = "Combat Modifiers"
mod_sheet['A1'].font = Font(bold=True, size=14)

mod_headers = ["Modifier_Type", "Condition", "Effect_Multiplier", "Description"]
for col, header in enumerate(mod_headers, start=1):
    cell = mod_sheet.cell(row=2, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

modifiers = [
    ["Terrain", "Urban", 0.7, "Reduced accuracy in urban terrain"],
    ["Terrain", "Forest", 0.8, "Reduced visibility in forests"],
    ["Terrain", "Open", 1.2, "Increased accuracy in open terrain"],
    ["Weather", "Rain", 0.85, "Reduced visibility and accuracy"],
    ["Weather", "Fog", 0.6, "Severe visibility reduction"],
    ["Time", "Night", 0.7, "Reduced visibility at night"],
    ["Status", "Damaged", 0.8, "Unit below 50% HP fights less effectively"],
    ["Status", "Suppressed", 0.5, "Unit under fire, accuracy reduced"],
    ["Flanking", "Side", 1.3, "Attacking from flank increases damage"],
    ["Flanking", "Rear", 1.5, "Attacking from rear bypasses armor"],
    ["Experience", "Veteran", 1.15, "Experienced crew bonus"],
    ["Experience", "Elite", 1.3, "Elite unit bonus"],
]

for row_data in modifiers:
    mod_sheet.append(row_data)

mod_sheet.column_dimensions['A'].width = 18
mod_sheet.column_dimensions['B'].width = 15
mod_sheet.column_dimensions['C'].width = 20
mod_sheet.column_dimensions['D'].width = 45

for row in range(3, mod_sheet.max_row + 1):
    mod_sheet.cell(row=row, column=3).number_format = '0.00'

# Add documentation sheet
doc_sheet = wb.create_sheet("Documentation")
doc_sheet['A1'] = "Combat Engagement System Documentation"
doc_sheet['A1'].font = Font(bold=True, size=14)

docs = [
    ["", ""],
    ["Field", "Description"],
    ["Attacker_Type", "Type of attacking unit"],
    ["Target_Type", "Type of target unit"],
    ["Base_Hit_Probability", "Base probability of hitting target (0.0-1.0)"],
    ["Damage_Multiplier", "Damage multiplier applied to Attack_Power"],
    ["Min_Range", "Minimum effective range in km"],
    ["Max_Range", "Maximum effective range in km"],
    ["Engagement_Priority", "Priority level (1=highest, 5=lowest)"],
    ["Notes", "Additional information about engagement type"],
    ["", ""],
    ["Damage Calculation", ""],
    ["Step 1", "Check if target is within range"],
    ["Step 2", "Calculate hit: random(0,1) < (Base_Hit_Probability * Attacker.Accuracy)"],
    ["Step 3", "If hit: Damage = Attacker.Attack_Power * Damage_Multiplier"],
    ["Step 4", "Apply armor: Final_Damage = max(0, Damage - Target.Armor * 0.5)"],
    ["Step 5", "Apply modifiers from Combat_Modifiers sheet if applicable"],
    ["Step 6", "Reduce Target.HP by Final_Damage"],
    ["", ""],
    ["Infantry Squad Composition", ""],
    ["Total", "10 personnel per squad"],
    ["7x", "Assault rifles (AK-74 or M4)"],
    ["1x", "RPG-7 or AT4 anti-tank"],
    ["1x", "Sniper rifle (SVD or M110)"],
    ["1x", "PKM or M249 machine gun"],
]

for row_idx, row_data in enumerate(docs, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = doc_sheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        if row_idx in [2, 12, 20]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

doc_sheet.column_dimensions['A'].width = 25
doc_sheet.column_dimensions['B'].width = 65

wb.save('/home/claude/combat_simulation/data/sets.xlsx')
print("sets.xlsx created successfully")
